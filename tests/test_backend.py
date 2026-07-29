"""后端逻辑测试。不依赖 PyQt，直接跑：python tests/test_backend.py"""

from __future__ import annotations

import sqlite3
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from backend import Backend, BackendError  # noqa: E402
from backend.constants import (  # noqa: E402
    DEVICE_TYPE_COLORS,
    DEVICE_TYPES,
    FALLBACK_DEVICE_TYPE,
)
from backend.models import (  # noqa: E402
    Cabinet,
    Device,
    DeviceLink,
    DeviceQuery,
    ImportOptions,
    RackRow,
    Reservation,
    Room,
)
from backend.schema import MIGRATIONS  # noqa: E402
from openpyxl import Workbook  # noqa: E402

PASSED: list[str] = []
FAILED: list[tuple[str, str]] = []


def check(name: str, condition: bool, detail: str = "") -> None:
    if condition:
        PASSED.append(name)
    else:
        FAILED.append((name, detail or "断言失败"))


def expect_error(name: str, fn, keyword: str = "") -> None:
    try:
        fn()
    except BackendError as exc:
        message = str(exc)
        if keyword and keyword not in message:
            FAILED.append((name, f"异常信息不含「{keyword}」：{message}"))
        else:
            PASSED.append(name)
    except Exception as exc:  # noqa: BLE001
        FAILED.append((name, f"抛出了非业务异常：{type(exc).__name__}: {exc}"))
    else:
        FAILED.append((name, "预期抛异常但没有"))


def new_device(name: str, **kwargs) -> Device:
    return Device(id=0, name=name, **kwargs)


def run(tmp: Path) -> None:
    api = Backend(tmp / "test.db")

    # ---------- 建库与示例数据 ----------
    check("库文件已创建", api.db_path.exists())
    stats = api.stats()
    check("示例数据：1 个机房", stats.rooms == 1, f"rooms={stats.rooms}")
    check("示例数据：5 个机柜", stats.cabinets == 5, f"cabinets={stats.cabinets}")
    check("示例数据：18 台设备", stats.devices == 18, f"devices={stats.devices}")
    check("示例数据：2 台待上架", stats.unracked == 2, f"unracked={stats.unracked}")
    check("示例数据：2 条预留", stats.reservations == 2, f"reservations={stats.reservations}")
    check("示例数据：1 台故障", stats.faulty == 1, f"faulty={stats.faulty}")

    cabinets = {c.name: c for c in api.list_cabinets()}
    a01, a02, a03 = cabinets["A01"], cabinets["A02"], cabinets["A03"]
    check("机柜带出机房名", a01.room_name == "示例机房", str(a01.room_name))
    check("机柜带出列名", a01.row_name == "A列", str(a01.row_name))

    # ---------- U 位冲突校验 ----------
    # A01 占用：40-41、37-38、34-35、1
    check("40U 已被占用", not api.check_placement(a01.id, 40, 1).ok)
    check("39U 空闲", api.check_placement(a01.id, 39, 1).ok)
    check("39U 放 2U 会压到 40U", not api.check_placement(a01.id, 39, 2).ok)
    out_of_range = api.check_placement(a01.id, 42, 2)
    check("42U 放 2U 越界", not out_of_range.ok and "超出机柜范围" in out_of_range.message,
          out_of_range.message)
    check("0U 非法", not api.check_placement(a01.id, 0, 1).ok)
    check("U 数为 0 非法", not api.check_placement(a01.id, 10, 0).ok)
    check("未指定 U 位视为未上架", api.check_placement(a01.id, None, 1).ok)
    check("预留区间同样占位", not api.check_placement(a02.id, 31, 1).ok)
    check("冲突信息里带占用者", "SW-CORE-01" in api.check_placement(a01.id, 40, 1).message)

    # ---------- 空闲区间 ----------
    slots = {(s.u_start, s.u_end) for s in api.free_slots(a01.id)}
    check("空闲区间含 2-33", (2, 33) in slots, str(sorted(slots)))
    check("空闲区间含 36-36", (36, 36) in slots, str(sorted(slots)))
    check("空闲区间含 39-39", (39, 39) in slots, str(sorted(slots)))
    check("空闲区间含 42-42", (42, 42) in slots, str(sorted(slots)))
    check("空闲区间不含被占的 40", not any(s <= 40 <= e for s, e in slots))

    # ---------- 设备增改 ----------
    created = api.save_device(
        new_device("TEST-SW-01", cabinet_id=a01.id, u_start=39, u_size=1,
                   dev_type="交换机", sn="SN-TEST-1", mgmt_ip="10.9.9.9", power_w=100, weight_kg=5)
    )
    check("新增设备成功", created.id > 0 and created.u_start == 39)
    check("新增后带出机柜名", created.cabinet_name == "A01", str(created.cabinet_name))
    check("u_end 计算正确", created.u_end == 39, str(created.u_end))

    expect_error(
        "新增到已占位置被拦截",
        lambda: api.save_device(new_device("TEST-DUP", cabinet_id=a01.id, u_start=40, u_size=1)),
        "U 位冲突",
    )
    expect_error("空设备名被拦截", lambda: api.save_device(new_device("   ")), "设备名")
    expect_error(
        "机柜不存在被拦截",
        lambda: api.save_device(new_device("TEST-NOCAB", cabinet_id=99999, u_start=1)),
        "机柜不存在",
    )

    created.mgmt_ip = "10.9.9.10"
    edited = api.save_device(created)
    check("编辑自身不与自己冲突", edited.mgmt_ip == "10.9.9.10")

    unknown_type = api.save_device(new_device("TEST-TYPE", dev_type="不存在的类型", status="乱写"))
    check("非法类型归为其他", unknown_type.dev_type == "其他", unknown_type.dev_type)
    check("非法状态归为在用", unknown_type.status == "在用", unknown_type.status)

    no_cab = api.save_device(new_device("TEST-NOCAB-U", cabinet_id=None, u_start=10))
    check("没机柜时 U 位被清空", no_cab.u_start is None)

    # ---------- 移动与上下架 ----------
    moved = api.move_device(created.id, a01.id, 20)
    check("移动到 20U", moved.u_start == 20)
    expect_error("移动到占用位被拦截", lambda: api.move_device(created.id, a01.id, 37), "U 位冲突")

    unracked = api.move_device(created.id, None, None)
    check("下架后机柜和 U 位都清空", unracked.cabinet_id is None and unracked.u_start is None)
    check("待上架列表包含它", any(d.id == created.id for d in api.list_unracked()))

    placed, failed = api.auto_rack([created.id], a01.id)
    check("自动上架选最低空位 2U", placed and placed[0][1] == 2, str(placed))
    check("自动上架无失败项", not failed, str(failed))

    # 1U 的小机柜放不下 4U 设备
    tiny_room = api.save_room(Room(id=0, name="测试机房"))
    tiny = api.save_cabinet(Cabinet(id=0, room_id=tiny_room.id, name="T01", u_total=1))
    big = api.save_device(new_device("TEST-BIG-4U", u_size=4, dev_type="服务器", status="备用"))
    placed2, failed2 = api.auto_rack([big.id], tiny.id)
    check("放不下时进 failed", not placed2 and len(failed2) == 1, str(failed2))
    check("失败原因说明清楚", "连续空位" in failed2[0][1], failed2[0][1])

    offline = api.save_device(new_device("TEST-OFFLINE", status="已下架", u_size=1))
    _, failed3 = api.auto_rack([offline.id], a03.id)
    check("已下架设备不自动上架", len(failed3) == 1 and "已下架" in failed3[0][1], str(failed3))

    # ---------- 机柜总高保护 ----------
    a01.u_total = 12
    expect_error("改小总高时保护已有内容", lambda: api.save_cabinet(a01), "无法把总高")
    a01.u_total = 42

    expect_error(
        "重复机柜编号被拦截",
        lambda: api.save_cabinet(Cabinet(id=0, room_id=a01.room_id, name="A01")),
        "已有编号",
    )
    expect_error(
        "重复机房名被拦截", lambda: api.save_room(Room(id=0, name="示例机房")), "已存在"
    )

    # ---------- 批量操作 ----------
    changed = api.bulk_update_devices([created.id, big.id], {"owner": "王工", "status": "备用"})
    check("批量修改影响 2 行", changed == 2, f"changed={changed}")
    check("批量修改结果正确", api.get_device(created.id).owner == "王工")
    expect_error(
        "批量修改校验状态取值",
        lambda: api.bulk_update_devices([created.id], {"status": "瞎写"}),
        "状态",
    )

    # ---------- 预留 ----------
    expect_error(
        "预留与设备冲突被拦截",
        lambda: api.save_reservation(Reservation(id=0, cabinet_id=a01.id, u_start=40, u_size=1)),
        "U 位冲突",
    )
    reservation = api.save_reservation(
        Reservation(id=0, cabinet_id=a01.id, u_start=25, u_size=3, label="测试预留")
    )
    check("预留创建成功", reservation.id > 0 and reservation.u_size == 3)
    check("预留 u_end 正确", reservation.u_end == 27)
    check("预留带出机柜名", reservation.cabinet_name == "A01", str(reservation.cabinet_name))
    check("预留后该位置不可用", not api.check_placement(a01.id, 26, 1).ok)
    api.delete_reservation(reservation.id)
    check("删除预留后位置释放", api.check_placement(a01.id, 26, 1).ok)

    # ---------- 机柜布局 ----------
    layout = api.cabinet_layout(a01.id)
    expected_used = sum(d.u_size for d in layout.racked_devices)
    check("布局 used_u 与设备累加一致", layout.used_u == expected_used,
          f"{layout.used_u} vs {expected_used}")
    check("布局 free_u 不为负", layout.free_u >= 0, str(layout.free_u))
    check(
        "布局三段相加等于总高",
        layout.used_u + layout.reserved_u + layout.free_u == layout.cabinet.u_total,
        f"{layout.used_u}+{layout.reserved_u}+{layout.free_u} != {layout.cabinet.u_total}",
    )
    check("布局功率累加正确", layout.power_used_w > 0, str(layout.power_used_w))

    # ---------- 容量汇总 ----------
    report = api.capacity_report()
    a01_row = next(r for r in report["cabinets"] if r.name == "A01")
    check("容量：A01 总 U 42", a01_row.u_total == 42)
    check(
        "容量：已用 + 预留 + 空闲 = 总 U",
        a01_row.u_used + a01_row.u_reserved + a01_row.u_free == 42,
        f"{a01_row.u_used}+{a01_row.u_reserved}+{a01_row.u_free}",
    )
    demo_room = next(r for r in report["rooms"] if r.name == "示例机房")
    check("容量：机房下 5 个机柜", demo_room.cabinet_count == 5, str(demo_room.cabinet_count))
    check("容量：机房总 U 210", demo_room.u_total == 210, str(demo_room.u_total))
    check("容量：占用率有值", demo_room.u_usage_pct > 0, str(demo_room.u_usage_pct))
    row_names = {r.name for r in report["rows"]}
    check("容量：列汇总含 A列 B列", {"A列", "B列"} <= row_names, str(row_names))
    check("容量：未分列单独成组", "未分列" in row_names, str(row_names))

    # 功率超限
    tiny.power_limit_w = 10
    api.save_cabinet(tiny)
    api.save_device(
        new_device("TEST-POWER-HOG", cabinet_id=tiny.id, u_start=1, u_size=1,
                   dev_type="服务器", power_w=500)
    )
    t01 = next(r for r in api.capacity_report()["cabinets"] if r.name == "T01")
    check("超限检测：功率超限", t01.overload, f"{t01.power_used_w}/{t01.power_limit_w}")
    check("超限检测：占用率计算", t01.power_usage_pct > 100, str(t01.power_usage_pct))

    tight = api.tightest_cabinets(3)
    check("最紧机柜里超限的排最前", tight[0].overload, tight[0].name)

    # ---------- 查询筛选 ----------
    check("关键字搜设备名", len(api.query_devices(DeviceQuery(keyword="CORE"))) >= 2)
    check("关键字搜 IP", any(d.mgmt_ip == "10.0.0.11"
                             for d in api.query_devices(DeviceQuery(keyword="10.0.0.11"))))
    check("关键字搜机柜名", len(api.query_devices(DeviceQuery(keyword="A02"))) >= 1)
    typed = api.query_devices(DeviceQuery(dev_types=("交换机",), statuses=("在用",)))
    check("类型 + 状态筛选", all(d.dev_type == "交换机" and d.status == "在用" for d in typed))
    by_cab = api.query_devices(DeviceQuery(cabinet_id=a01.id))
    check("按机柜筛选", all(d.cabinet_id == a01.id for d in by_cab))
    by_room = api.query_devices(DeviceQuery(room_id=a01.room_id))
    check("按机房筛选", len(by_room) > len(by_cab))
    paged = api.query_devices(DeviceQuery(sort_by="name", limit=5))
    check("排序分页", len(paged) == 5 and paged == sorted(paged, key=lambda d: d.name))
    total = api.count_devices(DeviceQuery())
    check("总数与查询一致", total == len(api.query_devices(DeviceQuery())), str(total))
    check("非法排序字段被忽略", len(api.query_devices(DeviceQuery(sort_by="drop table"))) > 0)
    check("厂商建议列表", "华为" in api.field_suggestions("vendor"))
    check("非白名单字段返回空", api.field_suggestions("password") == [])

    # ---------- 连接关系 ----------
    core1 = api.query_devices(DeviceQuery(keyword="SW-CORE-01"))[0]
    outgoing, incoming = api.list_links(core1.id)
    check("反向连接被统计", len(incoming) >= 3, f"incoming={len(incoming)}")
    acc1 = api.query_devices(DeviceQuery(keyword="SW-ACC-A02-1"))[0]
    acc_out, _ = api.list_links(acc1.id)
    check("上行连接解析出对端名", any(l.peer_resolved_name == "SW-CORE-01" for l in acc_out))

    text_link = api.save_link(
        DeviceLink(id=0, device_id=core1.id, local_port="XGE1/0/48",
                   peer_device_name="运营商设备", peer_port="GE0/1", speed="10G", medium="光纤")
    )
    check("对端可以只写文本", text_link.peer_device_name == "运营商设备")
    expect_error(
        "对端为空被拦截",
        lambda: api.save_link(DeviceLink(id=0, device_id=core1.id, local_port="X")),
        "对端设备不能为空",
    )
    expect_error(
        "对端不能是自己",
        lambda: api.save_link(DeviceLink(id=0, device_id=core1.id, peer_device_id=core1.id)),
        "不能是自己",
    )
    api.delete_link(text_link.id)
    after, _ = api.list_links(core1.id)
    check("删除连接生效", all(l.id != text_link.id for l in after))

    # 删除设备时连接级联清掉
    link_target = api.save_device(new_device("TEST-LINK-DEL"))
    api.save_link(DeviceLink(id=0, device_id=link_target.id, peer_device_name="X"))
    api.delete_devices([link_target.id])
    check("删设备后连接被级联删除",
          api.devices.links.count_for_device(link_target.id) == 0)

    # ---------- 删除机柜 / 机房时设备保留 ----------
    before = api.count_devices(DeviceQuery(unracked_only=True))
    moved_count = api.delete_cabinet(tiny.id)
    after_count = api.count_devices(DeviceQuery(unracked_only=True))
    check(
        "删机柜后设备转未上架而非删除",
        moved_count == 1 and after_count == before + 1,
        f"moved={moved_count} {before}->{after_count}",
    )

    room_to_delete = api.save_room(Room(id=0, name="待删机房"))
    cab_in_room = api.save_cabinet(Cabinet(id=0, room_id=room_to_delete.id, name="D01"))
    api.save_device(new_device("TEST-ROOM-DEL", cabinet_id=cab_in_room.id, u_start=1))
    total_before = api.count_devices(DeviceQuery())
    api.delete_room(room_to_delete.id)
    check("删机房后设备总数不变", api.count_devices(DeviceQuery()) == total_before)
    check("删机房级联删掉机柜", all(c.name != "D01" for c in api.list_cabinets()))

    # ---------- 批量建柜 ----------
    created_n, skipped = api.batch_create_cabinets(
        room_id=a01.room_id, row_id=a01.row_id, prefix="A", start_no=1, count=5,
        digits=2, u_total=42, power_limit_w=6000, weight_limit_kg=800,
    )
    check("批量建柜跳过已存在的", created_n == 2 and set(skipped) == {"A01", "A02", "A03"},
          f"created={created_n} skipped={skipped}")
    expect_error(
        "批量建柜数量上限",
        lambda: api.batch_create_cabinets(
            room_id=a01.room_id, row_id=None, prefix="X", start_no=1, count=999,
            digits=2, u_total=42,
        ),
        "数量",
    )
    check("批量建柜名称预览", api.preview_batch_names("B", 1, 10, 2) == ["B01", "B02", "B03", "B04"])

    # ---------- Excel ----------
    template = api.build_import_template(tmp / "template.xlsx")
    check("模板生成成功", template.exists())

    import_file = tmp / "import.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["机房", "列", "机柜编号", "起始U位", "占用U数", "设备名", "设备类型", "状态", "管理IP", "功耗W"])
    ws.append(["示例机房", "A列", "A03", 10, 2, "IMP-OK-1", "交换机", "在用", "10.5.5.1", 120])
    ws.append(["示例机房", "A列", "A03", 42, 1, "IMP-CONFLICT", "交换机", "在用", "10.5.5.2", 120])
    ws.append(["示例机房", "A列", "A03", 10, 1, "IMP-CROSSROW", "交换机", "在用", "10.5.5.3", 120])
    ws.append(["新机房X", "Z列", "Z09", 5, 1, "IMP-NEWPLACE", "路由器", "在用", "10.5.5.4", 90])
    ws.append(["示例机房", "A列", "A03", "", 1, "IMP-NOU", "交换机", "备用", "10.5.5.5", 60])
    ws.append(["示例机房", "A列", "A03", 200, 1, "IMP-OUTOFRANGE", "交换机", "在用", "", 60])
    wb.save(import_file)

    count_before = api.count_devices(DeviceQuery())
    dry = api.import_devices(import_file, ImportOptions(dry_run=True))
    check("预检不写库", api.count_devices(DeviceQuery()) == count_before,
          f"{count_before} -> {api.count_devices(DeviceQuery())}")
    check("预检统计 6 行", dry.total == 6, f"total={dry.total}")
    check("预检发现 3 处错误", len(dry.errors) == 3, str([(e.row, e.message) for e in dry.errors]))
    check("预检发现跨行冲突", any(e.row == 4 and "冲突" in e.message for e in dry.errors),
          str([(e.row, e.message) for e in dry.errors]))
    check("预检发现同柜已占位", any(e.row == 3 and "冲突" in e.message for e in dry.errors))
    check("预检发现越界", any("超出机柜范围" in e.message for e in dry.errors))
    check("预检可写入 3 行", dry.inserted == 3 and dry.writable == 3, f"inserted={dry.inserted}")

    real = api.import_devices(import_file, ImportOptions(dry_run=False))
    check("正式导入新增 3 行", real.inserted == 3, f"inserted={real.inserted}")
    check("导入后总数 +3", api.count_devices(DeviceQuery()) == count_before + 3)
    check("自动建了新机房", any(r.name == "新机房X" for r in api.list_rooms()))
    check("自动建了新机柜", any(c.name == "Z09" for c in api.list_cabinets()))
    imp_nou = api.query_devices(DeviceQuery(keyword="IMP-NOU"))[0]
    check("未填 U 位的设备算待上架", imp_nou.u_start is None and not imp_nou.is_racked)
    check("未填 U 位的设备出现在待上架列表",
          any(d.name == "IMP-NOU" for d in api.list_unracked()))
    check("导入保留了机柜归属", imp_nou.cabinet_name == "A03", str(imp_nou.cabinet_name))

    again = api.import_devices(import_file, ImportOptions(dry_run=False))
    check("重复导入走更新", again.updated == 3 and again.inserted == 0,
          f"updated={again.updated} inserted={again.inserted}")

    skip_run = api.import_devices(
        import_file, ImportOptions(dry_run=True, update_existing=False)
    )
    check("不更新模式记为跳过", skip_run.skipped == 3, f"skipped={skip_run.skipped}")

    # 用一个确实不存在的机房验证「不自动建柜」的行为
    missing_place_file = tmp / "missing_place.xlsx"
    wb_mp = Workbook()
    ws_mp = wb_mp.active
    ws_mp.append(["机房", "机柜编号", "设备名", "起始U位"])
    ws_mp.append(["从没见过的机房", "Q01", "IMP-NOPLACE", 5])
    ws_mp.append(["示例机房", "从没见过的机柜", "IMP-NOCAB", 5])
    wb_mp.save(missing_place_file)

    no_create = api.import_devices(
        missing_place_file, ImportOptions(dry_run=True, create_missing_places=False)
    )
    check("关掉自动建柜后报缺机房",
          any("机房「从没见过的机房」不存在" in e.message for e in no_create.errors),
          str([e.message for e in no_create.errors]))
    check("关掉自动建柜后报缺机柜",
          any("机柜「从没见过的机柜」" in e.message for e in no_create.errors),
          str([e.message for e in no_create.errors]))
    check("缺机房机柜时不写入任何行", no_create.inserted == 0, f"inserted={no_create.inserted}")

    # 别名表头
    alias_file = tmp / "alias.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["所属机房", "所在列", "机柜号", "U位置", "设备高度", "主机名", "设备种类", "ip地址"])
    ws2.append(["示例机房", "A列", "A03", 15, 1, "ALIAS-SW-01", "交换机", "10.6.6.1"])
    wb2.save(alias_file)
    alias_res = api.import_devices(alias_file, ImportOptions(dry_run=True))
    check("表头别名可识别", alias_res.inserted == 1 and not alias_res.errors,
          str([e.message for e in alias_res.errors]))

    # 缺必填列
    bad_file = tmp / "bad.xlsx"
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(["随便", "乱写", "的表头"])
    ws3.append(["a", "b", "c"])
    wb3.save(bad_file)
    bad_res = api.import_devices(bad_file, ImportOptions(dry_run=True))
    check("缺必填列给出提示",
          any("表头缺少必填列" in e.message for e in bad_res.errors),
          str([e.message for e in bad_res.errors]))

    # 空文件
    empty_file = tmp / "empty.xlsx"
    wb4 = Workbook()
    ws4 = wb4.active
    ws4.append(["机房", "机柜编号", "设备名"])
    wb4.save(empty_file)
    empty_res = api.import_devices(empty_file, ImportOptions(dry_run=True))
    check("空数据文件给出提示", any("没有数据行" in e.message for e in empty_res.errors))

    # 导出与往返
    dev_out = api.export_devices(tmp / "out_devices.xlsx")
    cap_out = api.export_capacity(tmp / "out_capacity.xlsx")
    check("设备清单导出成功", dev_out.exists())
    check("容量报表导出成功", cap_out.exists())

    round_trip = api.import_devices(
        dev_out, ImportOptions(dry_run=True, create_missing_places=False)
    )
    check(
        "导出文件可原样导回（往返一致）",
        not round_trip.errors,
        str([(e.row, e.message) for e in round_trip.errors][:5]),
    )
    check("往返全部是更新", round_trip.inserted == 0, f"inserted={round_trip.inserted}")

    filtered_out = api.export_devices(
        tmp / "out_filtered.xlsx", DeviceQuery(dev_types=("交换机",))
    )
    check("按筛选条件导出", filtered_out.exists())

    # ---------- 备份与切库 ----------
    backup = api.backup(tmp / "backup.db")
    check("备份文件生成", backup.exists() and backup.stat().st_size > 0)
    check("库大小可读", api.db_size_kb() > 0, str(api.db_size_kb()))

    other = Backend(tmp / "other.db", seed_demo=False)
    check("新库不铺示例数据", other.stats().rooms == 0)
    other.close()

    devices_before_switch = api.count_devices(DeviceQuery())
    api.switch_database(backup)
    check("切到备份库后数据一致",
          api.count_devices(DeviceQuery()) == devices_before_switch,
          f"{devices_before_switch} vs {api.count_devices(DeviceQuery())}")

    # ---------- 复制设备 ----------
    src = api.save_device(new_device("CP-SRV-01", dev_type="服务器", u_size=2,
                                     model="R740", vendor="Dell", sn="SN-CP-1",
                                     asset_no="ZC-CP-1", mgmt_ip="10.9.9.9",
                                     power_w=350.0, owner="张三", project="甲项目"))
    dup = api.copy_of_device(src.id)
    check("副本未落库（id 为 0）", dup.id == 0, str(dup.id))
    check("副本自动改名", dup.name == "CP-SRV-02", dup.name)
    check("型号厂商照抄", dup.model == "R740" and dup.vendor == "Dell",
          f"{dup.model}/{dup.vendor}")
    check("类型与 U 数照抄", dup.dev_type == "服务器" and dup.u_size == 2,
          f"{dup.dev_type}/{dup.u_size}")
    check("功耗责任人项目照抄",
          dup.power_w == 350.0 and dup.owner == "张三" and dup.project == "甲项目")
    check("SN 清空", not dup.sn, str(dup.sn))
    check("资产号清空", not dup.asset_no, str(dup.asset_no))
    check("管理 IP 清空", not dup.mgmt_ip, str(dup.mgmt_ip))
    check("U 位清空（避免与原机冲突）", dup.u_start is None, str(dup.u_start))

    saved_dup = api.save_device(dup)
    check("副本能直接存下", saved_dup.id > 0 and saved_dup.id != src.id)
    check("原设备未被改动", api.get_device(src.id).name == "CP-SRV-01")
    check("原设备 SN 仍在", api.get_device(src.id).sn == "SN-CP-1")

    # 名字占用了就继续往后找
    dup2 = api.copy_of_device(src.id)
    check("再复制跳过已占用编号", dup2.name == "CP-SRV-03", dup2.name)

    # 位数保持：09 -> 10，不是 010
    pdu = api.save_device(new_device("CP-PDU-09"))
    check("递增保持位数", api.copy_of_device(pdu.id).name == "CP-PDU-10",
          api.copy_of_device(pdu.id).name)

    # 末尾没数字就加副本，且不套娃
    fw = api.save_device(new_device("CP-防火墙"))
    first = api.save_device(api.copy_of_device(fw.id))
    check("无数字结尾加副本", first.name == "CP-防火墙-副本", first.name)
    check("副本不套娃", api.copy_of_device(first.id).name == "CP-防火墙-副本2",
          api.copy_of_device(first.id).name)

    expect_error("复制不存在的设备报错", lambda: api.copy_of_device(999999))

    # ---------- 设备类型自定义 ----------
    types = api.list_device_types()
    check("内置类型 11 种", len(types) == 11, str(len(types)))
    check("类型清单已同步到注册表",
          [t.name for t in types] == list(DEVICE_TYPES),
          f"{[t.name for t in types]} vs {list(DEVICE_TYPES)}")
    check("兜底类型在清单里", FALLBACK_DEVICE_TYPE in DEVICE_TYPES)
    check("内置类型标记为内置", all(t.builtin for t in types))

    added_type = api.create_device_type("动环监控", "#13c2c2")
    check("新增类型返回对象", added_type.name == "动环监控", added_type.name)
    check("新增类型不算内置", not added_type.builtin)
    check("新增后进注册表", "动环监控" in DEVICE_TYPES, str(DEVICE_TYPES))
    check("新增后配色进注册表",
          DEVICE_TYPE_COLORS.get("动环监控") == "#13c2c2",
          str(DEVICE_TYPE_COLORS.get("动环监控")))
    check("新类型排在兜底类型前",
          DEVICE_TYPES.index("动环监控") < DEVICE_TYPES.index(FALLBACK_DEVICE_TYPE),
          str(DEVICE_TYPES))

    # 自定义类型必须能真的存进台账，不能被归到「其他」
    env_dev = api.save_device(new_device("ENV-01", dev_type="动环监控"))
    check("设备可用自定义类型", env_dev.dev_type == "动环监控", env_dev.dev_type)
    check("自定义类型算进设备数",
          next(t.device_count for t in api.list_device_types() if t.name == "动环监控") == 1)

    # 改名要连带刷台账，否则设备的类型会变成清单外的悬空值
    renamed, moved = api.update_device_type("动环监控", "环境监控", "#08979c")
    check("改名返回新名", renamed.name == "环境监控", renamed.name)
    check("改名同步台账设备数", moved == 1, str(moved))
    check("改名后设备类型跟着改",
          api.get_device(env_dev.id).dev_type == "环境监控",
          api.get_device(env_dev.id).dev_type)
    check("旧名字从注册表移除", "动环监控" not in DEVICE_TYPES, str(DEVICE_TYPES))
    check("改名后配色同步", DEVICE_TYPE_COLORS.get("环境监控") == "#08979c")

    _, no_move = api.update_device_type("环境监控", "环境监控", "#eb2f96")
    check("只改色不动台账", no_move == 0, str(no_move))
    check("只改色配色生效", DEVICE_TYPE_COLORS.get("环境监控") == "#eb2f96")

    # 筛选认自定义类型
    check("按自定义类型筛选",
          len(api.query_devices(DeviceQuery(dev_types=("环境监控",)))) == 1)

    # 删类型：设备归兜底，不能连着设备一起删
    affected = api.delete_device_type("环境监控")
    check("删类型返回受影响设备数", affected == 1, str(affected))
    check("删类型后设备归兜底",
          api.get_device(env_dev.id).dev_type == FALLBACK_DEVICE_TYPE,
          api.get_device(env_dev.id).dev_type)
    check("删类型不删设备", api.get_device(env_dev.id).name == "ENV-01")
    check("删后从注册表移除", "环境监控" not in DEVICE_TYPES, str(DEVICE_TYPES))

    expect_error("兜底类型不能删",
                 lambda: api.delete_device_type(FALLBACK_DEVICE_TYPE), "兜底")
    expect_error("兜底类型不能改名",
                 lambda: api.update_device_type(FALLBACK_DEVICE_TYPE, "杂项", "#595959"),
                 "兜底")
    expect_error("类型不能重名", lambda: api.create_device_type("交换机"), "已经存在")
    expect_error("类型名不能为空", lambda: api.create_device_type("   "), "不能为空")
    expect_error("类型名不能过长",
                 lambda: api.create_device_type("超过十六个字符的类型名称肯定要被拦下来"),
                 "最多")
    expect_error("类型名不能含斜杠", lambda: api.create_device_type("交换机/路由器"), "「/」")
    expect_error("配色必须是十六进制", lambda: api.create_device_type("测试类型", "红色"), "#RRGGBB")
    expect_error("改不存在的类型报错",
                 lambda: api.update_device_type("不存在的类型", "新名", "#595959"), "不存在")
    expect_error("删不存在的类型报错", lambda: api.delete_device_type("不存在的类型"), "不存在")

    # 兜底类型改色是允许的
    api.update_device_type(FALLBACK_DEVICE_TYPE, FALLBACK_DEVICE_TYPE, "#8c8c8c")
    check("兜底类型可以改色", DEVICE_TYPE_COLORS.get(FALLBACK_DEVICE_TYPE) == "#8c8c8c")

    # 删掉内置类型也能恢复
    api.delete_device_type("KVM")
    check("内置类型可删", "KVM" not in DEVICE_TYPES, str(DEVICE_TYPES))
    restored = api.restore_default_device_types()
    check("恢复默认补回缺失项", restored == 1, str(restored))
    check("恢复后 KVM 回来了", "KVM" in DEVICE_TYPES, str(DEVICE_TYPES))
    check("恢复默认不重复补", api.restore_default_device_types() == 0)

    # 自定义类型不该因为恢复默认或清空数据消失
    api.create_device_type("自定义留存", "#722ed1")
    api.restore_default_device_types()
    check("恢复默认保留自定义类型", "自定义留存" in DEVICE_TYPES, str(DEVICE_TYPES))

    # 导入时认自定义类型，认不出的才归兜底
    type_xlsx = tmp / "custom_type.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["机房", "机柜编号", "设备名", "设备类型"])
    ws2.append(["示例机房", "A01", "IMP-CUSTOM", "自定义留存"])
    ws2.append(["示例机房", "A01", "IMP-UNKNOWN", "查无此类型"])
    wb2.save(type_xlsx)
    api.import_devices(type_xlsx, ImportOptions(dry_run=False, create_missing_places=True))
    imported = {d.name: d.dev_type for d in api.query_devices(DeviceQuery(keyword="IMP-"))}
    check("导入认自定义类型",
          imported.get("IMP-CUSTOM") == "自定义留存", str(imported))
    check("导入认不出的归兜底",
          imported.get("IMP-UNKNOWN") == FALLBACK_DEVICE_TYPE, str(imported))

    # ---------- 清空 ----------
    api.clear_all_data()
    check("清空数据保留类型配置",
          "自定义留存" in [t.name for t in api.list_device_types()],
          str([t.name for t in api.list_device_types()]))
    cleared = api.stats()
    check(
        "清空后所有表为空",
        cleared.rooms == 0 and cleared.cabinets == 0 and cleared.devices == 0
        and cleared.reservations == 0,
        str(cleared.to_dict()),
    )
    check("清空后仍可新增", api.save_room(Room(id=0, name="新机房")).id > 0)

    api.close()


def run_migration(tmp: Path) -> None:
    """老库（v1，没有 device_type 表）升级后类型清单要对得上。"""
    old = tmp / "legacy.db"
    conn = sqlite3.connect(str(old))
    conn.executescript(MIGRATIONS[0][1])
    conn.execute("PRAGMA user_version = 1")
    conn.execute("INSERT INTO device (name, dev_type, u_size) VALUES ('L-SW1','交换机',1)")
    # 手工改过库的情况：台账里有当年硬编码清单之外的类型
    conn.execute("INSERT INTO device (name, dev_type, u_size) VALUES ('L-TAPE','磁带库',2)")
    conn.commit()
    conn.close()

    api = Backend(old, seed_demo=False)
    names = [t.name for t in api.list_device_types()]
    check("老库升级后建出类型清单", len(names) >= 11, str(names))
    check("老库里清单外的类型被收进来", "磁带库" in names, str(names))
    check("升级不改台账里的类型",
          api.get_device(2).dev_type == "磁带库", api.get_device(2).dev_type)
    counts = {t.name: t.device_count for t in api.list_device_types()}
    check("升级后设备数统计正确",
          counts.get("交换机") == 1 and counts.get("磁带库") == 1, str(counts))

    # 切库要把注册表换成新库的清单，不能留着上一个库的自定义类型
    fresh = tmp / "fresh.db"
    api.switch_database(fresh)
    check("切库后注册表重灌", "磁带库" not in DEVICE_TYPES, str(DEVICE_TYPES))
    check("切库后内置类型在位", "交换机" in DEVICE_TYPES, str(DEVICE_TYPES))
    api.close()


def main() -> int:
    # Windows 上 WAL 附属文件偶尔来不及释放，清理失败不该影响结论
    with tempfile.TemporaryDirectory(
        prefix="cabinet-test-", ignore_cleanup_errors=True
    ) as folder:
        try:
            run(Path(folder))
            run_migration(Path(folder))
        except Exception as exc:  # noqa: BLE001
            import traceback

            print("测试中断：", exc)
            traceback.print_exc()
            FAILED.append(("测试执行", str(exc)))

    total = len(PASSED) + len(FAILED)
    print(f"\n用例 {total} 个，通过 {len(PASSED)}，失败 {len(FAILED)}")
    for name, detail in FAILED:
        print(f"  x {name} —— {detail}")
    if not FAILED:
        print("全部通过")
    return 1 if FAILED else 0


if __name__ == "__main__":
    raise SystemExit(main())
