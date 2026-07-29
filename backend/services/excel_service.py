"""Excel 模板、导入、导出。

导入分两步：预检在事务里真跑一遍再整体回滚，
所以同一批数据内部的 U 位冲突也能提前发现，不会导一半污染台账。
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..constants import DEVICE_STATUSES, DEVICE_TYPES
from ..database import Database
from ..errors import BackendError
from ..models import (
    Cabinet,
    Device,
    ImportError_,
    ImportOptions,
    ImportResult,
    RackRow,
    Room,
)
from ..repositories import DeviceRepository, PlaceRepository
from .capacity_service import CapacityService
from .occupancy import OccupancyService

# 模板列：(字段名, 表头, 列宽, 是否必填)
COLUMNS: tuple[tuple[str, str, int, bool], ...] = (
    ("room_name", "机房", 16, True),
    ("row_name", "列", 10, False),
    ("cabinet_name", "机柜编号", 16, True),
    ("u_start", "起始U位", 10, False),
    ("u_size", "占用U数", 10, False),
    ("name", "设备名", 22, True),
    ("dev_type", "设备类型", 12, False),
    ("status", "状态", 10, False),
    ("model", "型号", 20, False),
    ("vendor", "厂商", 14, False),
    ("sn", "序列号", 22, False),
    ("asset_no", "资产编号", 18, False),
    ("mgmt_ip", "管理IP", 16, False),
    ("power_w", "功耗W", 10, False),
    ("weight_kg", "重量kg", 10, False),
    ("install_date", "上架日期", 14, False),
    ("warranty_end", "保修到期", 14, False),
    ("owner", "责任人", 12, False),
    ("project", "项目/业务", 18, False),
    ("remark", "备注", 24, False),
)

# 常见表头别名，让用户现成的台账也能直接导
HEADER_ALIASES: dict[str, str] = {
    "机房": "room_name", "机房名称": "room_name", "数据中心": "room_name", "所属机房": "room_name",
    "列": "row_name", "列号": "row_name", "机柜列": "row_name", "所在列": "row_name",
    "机柜编号": "cabinet_name", "机柜": "cabinet_name", "机柜名称": "cabinet_name", "机柜号": "cabinet_name",
    "起始u位": "u_start", "起始u": "u_start", "u位": "u_start", "起始位置": "u_start", "u位置": "u_start",
    "占用u数": "u_size", "u数": "u_size", "高度u": "u_size", "设备高度": "u_size", "占用高度": "u_size",
    "设备名": "name", "设备名称": "name", "主机名": "name", "名称": "name", "设备": "name",
    "设备类型": "dev_type", "类型": "dev_type", "设备种类": "dev_type",
    "状态": "status", "使用状态": "status",
    "型号": "model", "设备型号": "model", "规格型号": "model",
    "厂商": "vendor", "品牌": "vendor", "厂家": "vendor", "生产厂商": "vendor",
    "序列号": "sn", "sn": "sn", "sn号": "sn", "序号": "sn",
    "资产编号": "asset_no", "资产号": "asset_no", "固资编号": "asset_no",
    "管理ip": "mgmt_ip", "ip": "mgmt_ip", "ip地址": "mgmt_ip", "管理地址": "mgmt_ip",
    "功耗w": "power_w", "功耗": "power_w", "额定功率": "power_w",
    "重量kg": "weight_kg", "重量": "weight_kg",
    "上架日期": "install_date", "安装日期": "install_date", "投产日期": "install_date",
    "保修到期": "warranty_end", "保修期至": "warranty_end", "过保日期": "warranty_end",
    "责任人": "owner", "负责人": "owner", "维护人": "owner",
    "项目/业务": "project", "项目": "project", "业务": "project", "所属业务": "project",
    "备注": "remark", "说明": "remark",
}

_HEADER_FILL = PatternFill("solid", fgColor="FFE6F4FF")
_HEADER_FONT = Font(bold=True)


def _norm_header(text: Any) -> str:
    return re.sub(r"[\s()（）*:：]", "", str(text or "")).lower()


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^\d\-]", "", str(value))
    return int(digits) if digits else None


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value))
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


def _as_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    match = re.match(r"^(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        y, m, d = match.groups()
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return None


class ExcelService:
    def __init__(self, db: Database) -> None:
        self.db = db
        self.places = PlaceRepository(db)
        self.devices = DeviceRepository(db)
        self.occupancy = OccupancyService(db)
        self.capacity = CapacityService(db)

    # ---------- 模板 ----------

    def build_template(self, target: str | Path) -> Path:
        target = Path(target)
        wb = Workbook()
        ws = wb.active
        ws.title = "设备台账"

        headers = [f"{h}*" if required else h for _, h, _, required in COLUMNS]
        ws.append(headers)
        for idx, (_, _, width, _) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
            cell = ws.cell(row=1, column=idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        ws.append(
            [
                "主机房", "A列", "A01", 40, 1, "SW-CORE-01", "交换机", "在用",
                "S12700E-4", "华为", "G1QW1234567", "ZC-2024-0001", "10.0.0.11",
                900, 45, "2024-03-15", "2027-03-14", "张工", "园区核心网",
                "示例行，导入前请删除",
            ]
        )

        help_ws = wb.create_sheet("填写说明")
        help_ws.column_dimensions["A"].width = 16
        help_ws.column_dimensions["B"].width = 88
        help_ws.append(["字段", "说明"])
        for cell in help_ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        notes = [
            ("带 * 的列", "必填。机房 + 机柜编号用来定位位置，设备名是识别依据之一。"),
            ("起始U位", "从机柜底部往上数，1 表示最底层。留空表示暂不上架，会进「待上架」列表。"),
            ("占用U数", "设备实际占用高度，留空按 1U 处理。"),
            ("设备类型", " / ".join(DEVICE_TYPES) + "。填其他值会归入「其他」。"),
            ("状态", " / ".join(DEVICE_STATUSES) + "。留空按「在用」处理，已下架的设备不占 U 位。"),
            ("日期列", "格式 2024-03-15，也支持 Excel 自带的日期格式。"),
            ("去重规则", "优先按序列号匹配，其次资产编号，都没有时按 机房 + 机柜 + 设备名 匹配。"),
            ("自动建柜", "导入时可勾选自动创建缺失的机房 / 列 / 机柜，新机柜的总 U 数用界面上设定的默认值。"),
            ("预检", "预检会在事务里真跑一遍再回滚，同一批数据内部的 U 位重叠也能提前发现。"),
            ("表头别名", "常见写法都能识别，比如「设备名称」「主机名」都会认成设备名。"),
        ]
        for item in notes:
            help_ws.append(list(item))
        for row in help_ws.iter_rows(min_row=2):
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

        target.parent.mkdir(parents=True, exist_ok=True)
        wb.save(target)
        return target

    # ---------- 读取 ----------

    def _read_rows(self, file: str | Path) -> tuple[list[dict[str, Any]], list[str]]:
        try:
            wb = load_workbook(file, data_only=True, read_only=True)
        except Exception as exc:  # openpyxl 抛的异常类型比较杂
            raise BackendError(f"打不开这个 Excel 文件：{exc}") from exc

        try:
            ws = next(
                (s for s in wb.worksheets if s.max_row and s.max_row > 1), wb.worksheets[0]
            )
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                return [], [h for _, h, _, req in COLUMNS if req]

            col_map: dict[int, str] = {}
            for index, cell in enumerate(header):
                field = HEADER_ALIASES.get(_norm_header(cell))
                if field and field not in col_map.values():
                    col_map[index] = field

            mapped = set(col_map.values())
            missing = [h for f, h, _, req in COLUMNS if req and f not in mapped]

            records: list[dict[str, Any]] = []
            for row_no, values in enumerate(rows_iter, start=2):
                record: dict[str, Any] = {"__row": row_no}
                has_value = False
                for index, field in col_map.items():
                    value = values[index] if index < len(values) else None
                    record[field] = value
                    if value not in (None, ""):
                        has_value = True
                if has_value:
                    records.append(record)
            return records, missing
        finally:
            wb.close()

    # ---------- 导入 ----------

    def import_devices(self, file: str | Path, options: ImportOptions) -> ImportResult:
        records, missing = self._read_rows(file)
        result = ImportResult(total=len(records))

        if missing:
            result.errors.append(
                ImportError_(
                    row=1,
                    message=f"表头缺少必填列：{'、'.join(missing)}。可以先下载模板对照列名。",
                )
            )
            return result
        if not records:
            result.errors.append(ImportError_(row=1, message="这个文件里没有数据行"))
            return result

        ctx = self.db.rollback_after() if options.dry_run else self.db.transaction()
        with ctx:
            self._import_rows(records, options, result)
        return result

    def _import_rows(
        self, records: list[dict[str, Any]], options: ImportOptions, result: ImportResult
    ) -> None:
        room_cache: dict[str, int] = {}
        row_cache: dict[tuple[int, str], int] = {}
        cab_cache: dict[tuple[int, str], int] = {}

        for record in records:
            row_no = int(record["__row"])
            try:
                self._import_one(record, row_no, options, result, room_cache, row_cache, cab_cache)
            except BackendError as exc:
                result.errors.append(ImportError_(row=row_no, message=str(exc)))
            except Exception as exc:  # 单行异常不该中断整批
                result.errors.append(ImportError_(row=row_no, message=f"处理失败：{exc}"))

    def _import_one(
        self,
        record: dict[str, Any],
        row_no: int,
        options: ImportOptions,
        result: ImportResult,
        room_cache: dict[str, int],
        row_cache: dict[tuple[int, str], int],
        cab_cache: dict[tuple[int, str], int],
    ) -> None:
        name = _as_text(record.get("name"))
        if not name:
            result.errors.append(ImportError_(row=row_no, field_name="设备名", message="设备名为空"))
            return

        room_name = _as_text(record.get("room_name"))
        cabinet_name = _as_text(record.get("cabinet_name"))
        u_start = _as_int(record.get("u_start"))
        u_size = max(1, _as_int(record.get("u_size")) or 1)

        # 机房和机柜都空，视为未上架设备。导出文件里的待上架行就是这样，保证往返可用
        if not room_name and not cabinet_name:
            self._upsert(record, row_no, None, None, u_size, name, options, result)
            return

        if not room_name:
            result.errors.append(ImportError_(row=row_no, field_name="机房", message="机房为空"))
            return
        if not cabinet_name:
            result.errors.append(
                ImportError_(row=row_no, field_name="机柜编号", message="机柜编号为空")
            )
            return

        room_id = room_cache.get(room_name)
        if room_id is None:
            room = self.places.find_room_by_name(room_name)
            if room:
                room_id = room.id
            elif options.create_missing_places:
                room_id = self.places.insert_room(Room(id=0, name=room_name))
            else:
                result.errors.append(
                    ImportError_(row=row_no, field_name="机房", message=f"机房「{room_name}」不存在")
                )
                return
            room_cache[room_name] = room_id

        row_id: int | None = None
        row_name = _as_text(record.get("row_name"))
        if row_name:
            cache_key = (room_id, row_name)
            row_id = row_cache.get(cache_key)
            if row_id is None:
                existing_row = self.places.find_row_by_name(room_id, row_name)
                if existing_row:
                    row_id = existing_row.id
                elif options.create_missing_places:
                    row_id = self.places.insert_row(
                        RackRow(id=0, room_id=room_id, name=row_name)
                    )
                if row_id:
                    row_cache[cache_key] = row_id

        cab_key = (room_id, cabinet_name)
        cabinet_id = cab_cache.get(cab_key)
        if cabinet_id is None:
            cabinet = self.places.find_cabinet_by_name(room_id, cabinet_name)
            if cabinet:
                cabinet_id = cabinet.id
            elif options.create_missing_places:
                cabinet_id = self.places.insert_cabinet(
                    Cabinet(
                        id=0,
                        room_id=room_id,
                        row_id=row_id,
                        name=cabinet_name,
                        u_total=options.default_u_total,
                    )
                )
            else:
                result.errors.append(
                    ImportError_(
                        row=row_no,
                        field_name="机柜编号",
                        message=f"机柜「{cabinet_name}」在机房「{room_name}」下不存在",
                    )
                )
                return
            cab_cache[cab_key] = cabinet_id

        self._upsert(record, row_no, cabinet_id, u_start, u_size, name, options, result)

    def _upsert(
        self,
        record: dict[str, Any],
        row_no: int,
        cabinet_id: int | None,
        u_start: int | None,
        u_size: int,
        name: str,
        options: ImportOptions,
        result: ImportResult,
    ) -> None:
        dev_type = _as_text(record.get("dev_type")) or "其他"
        if dev_type not in DEVICE_TYPES:
            dev_type = "其他"
        status = _as_text(record.get("status")) or "在用"
        if status not in DEVICE_STATUSES:
            status = "在用"

        sn = _as_text(record.get("sn"))
        asset_no = _as_text(record.get("asset_no"))

        existing = None
        if sn:
            existing = self.devices.find_by_sn(sn)
        if existing is None and asset_no:
            existing = self.devices.find_by_asset_no(asset_no)
        if existing is None:
            existing = (
                self.devices.find_by_cabinet_and_name(cabinet_id, name)
                if cabinet_id
                else self.devices.find_by_name(name)
            )

        if existing and not options.update_existing:
            result.skipped += 1
            return

        if cabinet_id and u_start is not None and status != "已下架":
            check = self.occupancy.check(
                cabinet_id,
                u_start,
                u_size,
                exclude_kind="device" if existing else "",
                exclude_id=existing.id if existing else 0,
            )
            if not check.ok:
                result.errors.append(
                    ImportError_(row=row_no, field_name="起始U位", message=check.message)
                )
                return

        device = Device(
            id=existing.id if existing else 0,
            cabinet_id=cabinet_id,
            name=name,
            u_start=u_start if cabinet_id else None,
            u_size=u_size,
            dev_type=dev_type,
            status=status,
            model=_as_text(record.get("model")),
            vendor=_as_text(record.get("vendor")),
            sn=sn,
            asset_no=asset_no,
            mgmt_ip=_as_text(record.get("mgmt_ip")),
            power_w=_as_float(record.get("power_w")),
            weight_kg=_as_float(record.get("weight_kg")),
            install_date=_as_date(record.get("install_date")),
            warranty_end=_as_date(record.get("warranty_end")),
            owner=_as_text(record.get("owner")),
            project=_as_text(record.get("project")),
            remark=_as_text(record.get("remark")),
        )

        if existing:
            self.devices.update(device)
            result.updated += 1
        else:
            self.devices.insert(device)
            result.inserted += 1

    # ---------- 导出 ----------

    def export_devices(self, target: str | Path, devices: list[Device]) -> Path:
        """导出成和导入模板一致的列结构，改完可以直接导回来。"""
        target = Path(target)
        wb = Workbook()
        ws = wb.active
        ws.title = "设备台账"

        ws.append([h for _, h, _, _ in COLUMNS])
        for idx, (_, _, width, _) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
            cell = ws.cell(row=1, column=idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        ws.freeze_panes = "A2"

        for d in devices:
            ws.append(
                [
                    d.room_name or "",
                    d.row_name or "",
                    d.cabinet_name or "",
                    d.u_start if d.u_start is not None else "",
                    d.u_size,
                    d.name,
                    d.dev_type,
                    d.status,
                    d.model or "",
                    d.vendor or "",
                    d.sn or "",
                    d.asset_no or "",
                    d.mgmt_ip or "",
                    d.power_w if d.power_w is not None else "",
                    d.weight_kg if d.weight_kg is not None else "",
                    d.install_date or "",
                    d.warranty_end or "",
                    d.owner or "",
                    d.project or "",
                    d.remark or "",
                ]
            )
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, len(devices) + 1)}"

        target.parent.mkdir(parents=True, exist_ok=True)
        wb.save(target)
        return target

    def export_capacity(self, target: str | Path) -> Path:
        target = Path(target)
        report = self.capacity.capacity_report()
        wb = Workbook()
        wb.remove(wb.active)

        headers = [
            ("名称", 18), ("所属", 16), ("机柜数", 9), ("设备数", 9),
            ("总U", 8), ("已用U", 9), ("预留U", 9), ("空闲U", 9), ("U占用率%", 11),
            ("功率上限W", 12), ("已用功率W", 12), ("承重上限kg", 12), ("已用承重kg", 12),
            ("超限", 8),
        ]

        for title, key in (("机房汇总", "rooms"), ("列汇总", "rows"), ("机柜明细", "cabinets")):
            ws = wb.create_sheet(title)
            ws.append([h for h, _ in headers])
            for idx, (_, width) in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
                cell = ws.cell(row=1, column=idx)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
            ws.freeze_panes = "A2"

            for r in report[key]:
                ws.append(
                    [
                        r.name, r.parent_name or "", r.cabinet_count, r.device_count,
                        r.u_total, r.u_used, r.u_reserved, r.u_free, r.u_usage_pct,
                        r.power_limit_w, r.power_used_w, r.weight_limit_kg, r.weight_used_kg,
                        "是" if r.overload else "",
                    ]
                )
                if r.overload:
                    for cell in ws[ws.max_row]:
                        cell.font = Font(color="FFCF1322", bold=True)

        target.parent.mkdir(parents=True, exist_ok=True)
        wb.save(target)
        return target
